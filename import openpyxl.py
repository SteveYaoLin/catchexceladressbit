import openpyxl
import os

def extract_signal_information(filepath, sheet_number):
    workbook = openpyxl.load_workbook(filepath)
    sheet_names = workbook.sheetnames

    if sheet_number < 1 or sheet_number > len(sheet_names):
        raise ValueError("Invalid sheet number")

    sheet = workbook[sheet_names[sheet_number - 1]]

    # 读取非零元素的信息
    signal_info = []
    offset_info = []
    bit_info = []

    for row in range(2, 6):  # 假设地址偏移量在第2行到第5行之间
        for col in range(2, 34):  # 假设bit数在第2列到第33列之间
            signal = sheet.cell(row=row, column=col).value
            offset = sheet.cell(row=row, column=1).value
            bit = sheet.cell(row=1, column=col).value

            # 跳过空单元格和标题行
            if signal is None or signal == "SIGNAL" or bit is None:
                continue

            signal_info.append(signal)
            offset_info.append(offset)
            bit_info.append(bit)

    return signal_info, offset_info, bit_info

def save_information_to_txt(signal_info, offset_info, bit_info, output_filepath):
    with open(output_filepath, 'w') as output_file:
        for signal, offset, bit in zip(signal_info, offset_info, bit_info):
            output_file.write(f"{signal}\t\t{offset}\t\t{bit}\n")

    print(f"信息已保存到：{output_filepath}")

    # 自动打开生成的 output.txt 文件
    try:
        os.startfile(output_filepath)
    except AttributeError:
        try:
            subprocess.call(['open', output_filepath])
        except:
            pass

# 获取用户输入的Excel文件路径
excel_filepath = input("请输入Excel文件路径：")

# 获取用户输入的sheet编号
sheet_number = int(input("请输入要处理的sheet编号："))

# 提取信息
signal_info, offset_info, bit_info = extract_signal_information(excel_filepath, sheet_number)

# 指定输出文件路径
output_filepath = 'output.txt'

# 保存信息到TXT文件
save_information_to_txt(signal_info, offset_info, bit_info, output_filepath)